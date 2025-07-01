import time
import os
import random
import sys
from selenium import webdriver
from selenium.common import NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from Lib.ExImages import FetchingImages
from Lib.Products import Actions, Categories
from Lib.Actions import AnalogWatches_Action as Action, Notifyer as Notify, ManageCache as Cache
from openpyxl import load_workbook



# Files Path
ProImgDir = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\g4+w2+s2+cp\\images'
SlideImagesPath = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\g4+w2+s2+cp\\slide'
path = 'C:\\Users\\bhatt\\Downloads\\Telegram Desktop\\g4+w2+s2+cp'
category = Categories.AnalogWatches
Slide_Images = True

options = webdriver.ChromeOptions()
# options.add_experimental_option('excludeSwitches', ['enable-logging'])


#driver = webdriver.Chrome(options=options)
driver = webdriver.Chrome(executable_path='C:\\chromedriver\\chromedriver.exe')


wb = load_workbook('Products.xlsx')
ws = wb['Entries']
ws2 = wb['srno']


def Restart():
    driver.close()
    os.execv(sys.executable, ['python'] + sys.argv)


def UploadProducts(images, driver):
    row = 1
    col = 1
    ibtn = 0
    for x in images:
        try:

            OverAllTimeStarts = time.time()
            print('-----------------------------------------------------------------------------------')



            a = ws2['A1']
            n = int(a.value)

            #Cache.ClearCache(driver)
            print("Email Account : " + Actions.email_AC)

            AddSingleCatlog = driver.find_element(By.XPATH,
                                                  '//*[@id="mainWrapper"]/div/div[1]/div[2]/div[1]/div[1]/div/button[2]')
            AddSingleCatlog.click()


            try:

                driver.implicitly_wait(7)
                Categories.Category(driver, Categories.Category(driver, category))
            except Exception as e:
                driver.refresh()
                driver.implicitly_wait(10)
                Categories.Category(driver, Categories.Category(driver, category))

            # Upload Product Front Image
            UploadProdImage = driver.find_element(By.XPATH, '//*[@id="addImages"]').send_keys(x)

            # time.sleep(5)
            driver.implicitly_wait(30)

            # Clicking to Continue button
            Continue = "//div[@role='dialog']/div/div[3]/button/span[contains(text(), 'Continue')]"
            ContinueBtn = driver.find_element(By.XPATH, Continue)
            ContinueBtn.click()

            # time.sleep(8)
            driver.implicitly_wait(15)


            try:
                idiv = "//span[normalize-space()='Ok, Understood']//parent::button"
                driver.implicitly_wait(5)

                if driver.find_element(By.XPATH, idiv).is_enabled():
                    driver.find_element(By.XPATH,'//p[normalize-space()="Wrong/Defective Returns Price"]//parent::div//parent::div/div[2]/div/div/div/div').click()
                    print("I-Button:True")


            except Exception as e:
                    print("I-Button:False")

            # # i Button
            # if ibtn <= 4:
            #     driver.implicitly_wait(15)
            #     driver.find_element(By.XPATH,
            #                         '//p[normalize-space()="Wrong/Defective Returns Price"]//parent::div//parent::div/div[2]/div/div/div/div').click()

            time.sleep(2)

            print(str(x) + ' ' + str(n))

            driver.execute_script("window.scrollBy(0, 200)", "")

            driver.implicitly_wait(20)
            UploadSlideImages = driver.find_element(By.XPATH, '//input[@id="addMoreImagesInput"]')

            if Slide_Images == True:
                # Upload Product Slide Image
                SlideImages = FetchingImages.SlideImages(SlideImagesPath)
                for s in SlideImages:
                    UploadSlideImages.send_keys(s)
                    time.sleep(1)

            char1 = random.choice(range(65, 90))
            char2 = random.choice(range(65, 90))

            alpha = chr(char1)
            # ------------------------------------------------------------------------------------------------------
            start = time.time()
            Action.AnalogWatches_InventoryDetails(driver, alpha)
            end = time.time()
            t = end - start
            print("TIme Taken : " + str(t))
            # ------------------------------------------------------------------------------------------------------


            # ------------------------------------------------------------------------------------------------------
            start = time.time()
            Action.AnalogWatches_ProductDetails(driver)
            end = time.time()
            t = end - start
            print("Time Taken : " + str(t))
            # ------------------------------------------------------------------------------------------------------


            # ------------------------------------------------------------------------------------------------------
            start = time.time()
            Action.AnalogWatches_OtherAttributes(driver)
            end = time.time()
            t = end - start
            print("Time Taken : " + str(t))
            # ------------------------------------------------------------------------------------------------------


            ws.cell(row=row, column=col).value = str(x) + ' ' + str(n)
            row += 1
            ibtn += 1

            SubmitButton = "//div[@data-testid='singleCatalogProductAdd']/div/div/button/span[contains(text(), 'Submit')]//parent::button"
            driver.find_element(By.XPATH, SubmitButton).click()
            time.sleep(1)
            driver.find_element(By.XPATH, "//div[@role='dialog']/div/div[2]/button/span[contains(text(), 'Proceed')]//parent::button").click()

            driver.implicitly_wait(15)
            print('Done')
            n += 1
            ws2['A1'] = n
            destination = path + '\\' + os.path.basename(x)
            os.replace(x, destination)
            print('File Name : (' + os.path.basename(x) + ')\n has been moved to destination : ' + destination)
            wb.save('Products.xlsx')
            OverAllTimeEnds = time.time()
            OverAllTime = OverAllTimeEnds - OverAllTimeStarts
            print('-----------------------------------------------------------------------------------')
            print("OverAll Time Taken : " + str(OverAllTime))
            print('-----------------------------------------------------------------------------------')
            print("\n")



        except (ElementClickInterceptedException, NoSuchElementException, StaleElementReferenceException) as e:
            print("Exception : " + str(e))
            Notify.ExceptionCatched(str(e))
            discardbtn = "//div[@data-testid='singleCatalogProductAdd']/div/button/span[contains(text(), 'Discard')]//parent::button"
            process = '//div[@class="css-1qehpk6"]//div[2]//button/span[contains(text(), "Proceed")]'

            wait = WebDriverWait(driver, 10, ignored_exceptions=[ElementClickInterceptedException])
            wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body')))
            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body')))
            driver.find_element(By.XPATH, '/html/body').click()

            wait = WebDriverWait(driver, 10, ignored_exceptions=[ElementClickInterceptedException])
            wait.until(EC.visibility_of_element_located((By.XPATH, discardbtn)))
            wait.until(EC.element_to_be_clickable((By.XPATH, discardbtn)))
            driver.find_element(By.XPATH, discardbtn).click()

            time.sleep(1)
            wait = WebDriverWait(driver, 10, ignored_exceptions=[ElementClickInterceptedException])
            wait.until(EC.visibility_of_element_located((By.XPATH, process)))
            wait.until(EC.element_to_be_clickable((By.XPATH, process)))
            driver.find_element(By.XPATH, process).click()

            driver.implicitly_wait(25)
            continue


if __name__ == "__main__":


    driver.get('https://supplier.meesho.com/')
    driver.maximize_window()

    start = time.time()
    Actions.Login(driver)

    Cache.PreparingCache(driver)
    end = time.time()
    t = end-start
    print("Time Taken : "+str(t))

    images = FetchingImages.ProductImages(ProImgDir)

    UploadProducts(images, driver)

    images2 = FetchingImages.ProductImages(ProImgDir)

    if len(images2) > 0:
        print("Remaning Images : "+str(len(images2)))
        UploadProducts(images2, driver)
    else:
        print("Remaning Images : "+str(len(images2)))
        ws2['A1'] = 1
        print('Out of images')
        wb.save('Products.xlsx')
        Notify.UploadeSucessfully()







