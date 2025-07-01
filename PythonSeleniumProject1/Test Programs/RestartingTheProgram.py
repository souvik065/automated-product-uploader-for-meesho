import openpyxl
from openpyxl import Workbook, load_workbook

wb = load_workbook('..\LearningSelenium\Products.xlsx')
ws = wb['srno']

n = ws['A1']
a = int(n.value)


print(a)
a += 1
ws['A1'] = a


n = ws['A1']
a = n.value
print(a)


wb.save('..\LearningSelenium\Products.xlsx')

